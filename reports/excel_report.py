# C:\Users\USER\Documents\개발 폴더\kvan-dashboard\reports\excel_report.py
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker

today_str = date.today().strftime("%Y-%m-%d")


def build_monthly_report(df, vendors, start_month, end_month):
    wb = Workbook()

    # =========================================================
    # 공통 스타일
    # =========================================================
    NAVY = "1F2A44"       # 메인 포인트 컬러
    LIGHT_GRAY = "F8F9FA" # 배경색
    BORDER_COLOR = "E5E7EB"
    WHITE = "FFFFFF"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR)
    )

    # =========================================================
    # 1️⃣ Dashboard 시트 생성 및 배경 설정
    # =========================================================
    ws = wb.active
    ws.title = "Dashboard"

    # 전체 배경색 도포 (A1:Z100)
    for r in range(1, 101):
        for c in range(1, 27):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # -------------------------
    # 메인 제목 영역
    # -------------------------
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = "해외부 매출 분석 리포트"
    title_cell.font = Font(bold=True, size=24, color=NAVY)
    title_cell.alignment = center

    ws.merge_cells("A3:H3")
    subtitle_cell = ws["A3"]
    subtitle_cell.value = f"조회 기간: {start_month} ~ {end_month}  |  출력일: {today_str}"
    subtitle_cell.font = Font(size=11, color="666666")
    subtitle_cell.alignment = center

    # -------------------------
    # KPI 섹션 (A5:H8)
    # -------------------------
    total_gross = df["gross_sales"].sum()
    total_net = df["net_sales"].sum()
    total_fee = df["vendor_fee"].sum()
    total_rides = int(df["ride_count"].sum())

    kpis = [
        ("총 매출액", f"₩ {total_gross:,.0f}"),
        ("실 입금액", f"₩ {total_net:,.0f}"),
        ("총 수수료", f"₩ {total_fee:,.0f}"),
        ("운행 건수", f"{total_rides:,} 건"),
    ]

    kpi_cols = [("A", "B"), ("C", "D"), ("E", "F"), ("G", "H")]

    for i, (title, value) in enumerate(kpis):
        start_col, end_col = kpi_cols[i]
        target_range = f"{start_col}5:{end_col}8"
        ws.merge_cells(target_range)
        
        cell = ws[f"{start_col}5"]
        # 리치 텍스트 효과를 위해 줄바꿈 사용
        cell.value = f"{title}\n\n{value}"
        cell.font = Font(bold=True, size=14, color=NAVY)
        cell.alignment = center
        cell.fill = PatternFill("solid", fgColor=WHITE)
        cell.border = thin_border

    # -------------------------
    # 2️⃣ 업체별 매출 분석 데이터 정리 (차트용)
    # -------------------------
    # 차트용 데이터를 시트의 구석(Z열)으로 숨겨서 배치 (Dashboard를 깔끔하게 유지)
    vendor_total = df.groupby("vendor", as_index=False).agg({"gross_sales": "sum"})
    
    data_start_row = 100 # 데이터는 보이지 않는 곳에 저장
    ws.cell(row=data_start_row, column=25, value="업체")
    ws.cell(row=data_start_row, column=26, value="매출")
    
    for i, row in vendor_total.iterrows():
        ws.cell(row=data_start_row + i + 1, column=25, value=row["vendor"])
        ws.cell(row=data_start_row + i + 1, column=26, value=row["gross_sales"])

    data_ref = Reference(ws, min_col=26, min_row=data_start_row + 1, max_row=data_start_row + len(vendor_total))
    cats_ref = Reference(ws, min_col=25, min_row=data_start_row + 1, max_row=data_start_row + len(vendor_total))

    # -------------------------
    # Bar Chart: 업체별 매출 비교
    # -------------------------
    bar = BarChart()
    bar.title = "업체별 매출 비교 (KRW)"
    bar.style = 10
    bar.width = 15
    bar.height = 9
    bar.legend = None
    bar.y_axis.majorGridlines = None # 그리드 제거로 깔끔하게

    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(cats_ref)
    
    # 데이터 레이블 설정 (막대 위 숫자)
    bar.dataLabels = DataLabelList()
    bar.dataLabels.showVal = True
    bar.dataLabels.numFmt = "#,##0" # 숫자 콤마 표시
    
    ws.add_chart(bar, "A10")

    # -------------------------
    # Pie Chart: 업체별 매출 비중
    # -------------------------
    pie = PieChart()
    pie.title = "업체별 매출 점유율"
    pie.width = 11
    pie.height = 9
    
    pie.add_data(data_ref, titles_from_data=False)
    pie.set_categories(cats_ref)
    
    # 레이블 설정 (업체명 + 백분율)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True # 업체명
    pie.dataLabels.showPercent = True # 퍼센트
    pie.dataLabels.separator = " : "
    
    # 범례 위치를 아래로 내려서 차트 영역 확보
    pie.legend.position = 'b'

    ws.add_chart(pie, "E10")

    # -------------------------
    # 3️⃣ 월별 매출 추이 (Line Chart)
    # -------------------------
    monthly_sum = df.groupby("month", as_index=False).agg({"gross_sales": "sum"}).sort_values("month")
    
    line_data_row = 120
    for i, row in monthly_sum.iterrows():
        ws.cell(row=line_data_row + i, column=25, value=row["month"])
        ws.cell(row=line_data_row + i, column=26, value=row["gross_sales"])

    l_data = Reference(ws, min_col=26, min_row=line_data_row, max_row=line_data_row + len(monthly_sum) - 1)
    l_cats = Reference(ws, min_col=25, min_row=line_data_row, max_row=line_data_row + len(monthly_sum) - 1)

    line = LineChart()
    line.title = "월별 매출액 추이"
    line.width = 27 # 가로로 길게 배치
    line.height = 10
    line.legend = None
    line.style = 13
    line.smooth = True # 곡선 처리
    
    line.add_data(l_data, titles_from_data=False)
    line.set_categories(l_cats)

    # 마커 및 레이블 설정
    line.dataLabels = DataLabelList()
    line.dataLabels.showVal = True
    line.dataLabels.numFmt = "#,##0"
    
    s1 = line.series[0]
    s1.marker = Marker(symbol='circle', size=7)
    s1.graphicalProperties.line.width = 30000 # 선 두께 조절 (EMU 단위)

    ws.add_chart(line, "A23")

    # -------------------------
    # 컬럼 너비 최적화
    # -------------------------
    for col in ["A","B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 16


    # =========================================================
    # 3️⃣ 시트 : 업체별 월매출 (🔥 완전 수정 🔥)
    # =========================================================
    ws3 = wb.create_sheet(title="업체별 월매출")

    ws3.merge_cells("A1:M1")
    ws3["A1"] = "해외부 월별 업체 매출"
    ws3["A1"].font = Font(bold=True, size=18)
    ws3["A1"].alignment = center

    ws3.merge_cells("A2:M2")
    ws3["A2"] = f"업체: {', '.join(vendors)} | 기간: {start_month} ~ {end_month}"
    ws3["A2"].alignment = center

    ws3["A3"] = f"작성일: {today_str}"
    ws3["A4"] = "담당자: 이수민"

    current_row = 6
    months = sorted(df["month"].unique())

    # --- 헤더 (한 번만)
    headers = ["업체", "구분"] + months + ["합계"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws3.cell(row=current_row, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = soft_border


    current_row += 1

    metrics = [
        ("매출액", "gross_sales"),
        ("업체 수수료", "vendor_fee"),
        ("실 입금액", "net_sales"),
        ("운행건수", "ride_count"),
    ]

    for vendor in vendors:
        vendor_df = df[df["vendor"] == vendor]
        start_vendor_row = current_row

        for label, col in metrics:
            ws3.cell(row=current_row, column=2, value=label).alignment = center
            ws3.cell(row=current_row, column=2).border = soft_border


            row_sum = 0
            for i, m in enumerate(months, start=3):
                v = vendor_df[vendor_df["month"] == m][col].sum()
                c = ws3.cell(row=current_row, column=i, value=v)
                c.border = soft_border
                c.alignment = center
                if col != "ride_count":
                    c.number_format = "#,##0"
                row_sum += v

            total_col = len(months) + 3
            c = ws3.cell(row=current_row, column=total_col, value=row_sum)
            c.font = bold_font
            c.border = soft_border
            c.alignment = center
            if col != "ride_count":
                c.number_format = "#,##0"

            current_row += 1

        # 업체명 세로 병합 (A열)
        ws3.merge_cells(
            start_row=start_vendor_row,
            start_column=1,
            end_row=current_row - 1,
            end_column=1
        )
        c = ws3.cell(row=start_vendor_row, column=1, value=vendor)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = soft_border

        current_row += 1  # 업체 간 여백

    # =========================
    # 🔥 총계 블록 (모든 업체 합산)
    # =========================
    total_start_row = current_row

    for label, col in metrics:
        ws3.cell(row=current_row, column=2, value=label).alignment = center
        ws3.cell(row=current_row, column=2).border = soft_border


        row_sum = 0
        for i, m in enumerate(months, start=3):
            v = df[df["month"] == m][col].sum()
            c = ws3.cell(row=current_row, column=i, value=v)
            c.border = soft_border
            c.alignment = center
            if col != "ride_count":
                c.number_format = "#,##0"
            row_sum += v

        total_col = len(months) + 3
        c = ws3.cell(row=current_row, column=total_col, value=row_sum)
        c.font = bold_font
        c.border = soft_border
        c.alignment = center
        if col != "ride_count":
            c.number_format = "#,##0"

        current_row += 1

    ws3.merge_cells(
        start_row=total_start_row,
        start_column=1,
        end_row=current_row - 1,
        end_column=1
    )
    c = ws3.cell(row=total_start_row, column=1, value="총계")
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = soft_border


    # 컬럼 너비
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 14
    for i in range(3, len(months) + 4):
        ws3.column_dimensions[get_column_letter(i)].width = 18

    # =========================
    # 저장
    # =========================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
